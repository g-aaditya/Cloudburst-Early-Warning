"""
Excel Data Loader
Reads weather parameters from cloudburst_data.xlsx
and writes prediction results back to the Prediction_Log sheet.
"""

import os
import openpyxl
import pandas as pd
from datetime import datetime
from dataclasses import dataclass
from typing import Optional, Tuple


EXCEL_FILE = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "cloudburst_data.xlsx")

# Cell map: parameter name → (sheet, row, col)
CELL_MAP = {
    "max_temp":           ("Weather_Input", 6,  2),
    "min_temp":           ("Weather_Input", 7,  2),
    "rainfall":           ("Weather_Input", 8,  2),
    "wind_speed":         ("Weather_Input", 9,  2),
    "wind_gust":          ("Weather_Input", 10, 2),
    "wind_direction":     ("Weather_Input", 11, 2),
    "humidity":           ("Weather_Input", 12, 2),
    "pressure":           ("Weather_Input", 13, 2),
    "elevation":          ("Weather_Input", 16, 2),
    "slope":              ("Weather_Input", 17, 2),
    "soil_moisture":      ("Weather_Input", 18, 2),
    "weather_code":       ("Weather_Input", 21, 2),
    "region":             ("Weather_Input", 22, 2),
    "population_density": ("Weather_Input", 23, 2),
}

DEFAULTS = {
    "max_temp": 32.0, "min_temp": 22.0, "rainfall": 40.0,
    "wind_speed": 25.0, "wind_gust": 40.0, "wind_direction": 210.0,
    "humidity": 75.0, "pressure": 995.0, "elevation": 1200.0,
    "slope": 22.0, "soil_moisture": 60.0, "weather_code": 63,
    "region": "himalayan", "population_density": "semi",
}

VALID_REGIONS = ["himalayan", "western_ghats", "northeast", "coastal", "plains"]
VALID_POPDENSITY = ["rural", "semi", "urban"]
VALID_WMO = [0, 1, 2, 3, 51, 61, 63, 65, 80, 82, 95, 99]


def get_excel_path() -> str:
    return EXCEL_FILE


def excel_exists() -> bool:
    return os.path.isfile(EXCEL_FILE)


def get_file_mtime() -> Optional[float]:
    if excel_exists():
        return os.path.getmtime(EXCEL_FILE)
    return None


def _coerce(key: str, raw):
    """Coerce and validate a raw cell value into the expected type.
    Out-of-range numeric values are CLAMPED (not rejected) so one bad
    cell never silently blocks the whole reload.
    """
    if raw is None:
        return DEFAULTS[key]

    if key in ("region", "population_density"):
        val = str(raw).strip().lower()
        if key == "region" and val not in VALID_REGIONS:
            return DEFAULTS[key]
        if key == "population_density" and val not in VALID_POPDENSITY:
            return DEFAULTS[key]
        return val

    if key == "weather_code":
        try:
            code = int(float(raw))
            # Find closest valid WMO code
            if code in VALID_WMO:
                return code
            # Clamp to nearest valid code instead of rejecting
            return min(VALID_WMO, key=lambda x: abs(x - code))
        except (ValueError, TypeError):
            return DEFAULTS[key]

    try:
        val = float(raw)
        # Range guards — CLAMP, don't reject
        guards = {
            "max_temp": (5, 45), "min_temp": (0, 35), "rainfall": (0, 200),
            "wind_speed": (0, 120), "wind_gust": (0, 180), "wind_direction": (0, 360),
            "humidity": (20, 100), "pressure": (950, 1030),
            "elevation": (100, 4500), "slope": (0, 60), "soil_moisture": (0, 100),
        }
        if key in guards:
            lo, hi = guards[key]
            val = max(lo, min(hi, val))   # clamp — never reject
        return val
    except (ValueError, TypeError):
        return DEFAULTS[key]


def load_from_excel() -> Tuple[dict, str]:
    """
    Read all parameters from the Excel file.
    Returns (params_dict, status_message).
    """
    if not excel_exists():
        return DEFAULTS.copy(), f"⚠ Excel file not found at: {EXCEL_FILE}\nUsing default values."

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True, read_only=False)
        params = {}
        errors = []

        for key, (sheet_name, row, col) in CELL_MAP.items():
            if sheet_name not in wb.sheetnames:
                params[key] = DEFAULTS[key]
                errors.append(f"Sheet '{sheet_name}' not found")
                continue
            ws = wb[sheet_name]
            raw = ws.cell(row=row, column=col).value
            coerced = _coerce(key, raw)
            params[key] = coerced

        wb.close()

        ts = datetime.now().strftime("%H:%M:%S")
        if errors:
            msg = f"⚠ Loaded with warnings at {ts}:\n" + "\n".join(errors[:3])
        else:
            msg = f"✅ Loaded from Excel at {ts}"
        return params, msg

    except Exception as e:
        return DEFAULTS.copy(), f"❌ Error reading Excel: {e}\nUsing defaults."


def load_historical_data() -> Optional[pd.DataFrame]:
    """Read the Historical_Data sheet as a DataFrame."""
    if not excel_exists():
        return None
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name="Historical_Data", header=2)
        df.columns = df.columns.str.strip()
        return df
    except Exception:
        return None


def append_prediction_log(result_dict: dict) -> bool:
    """
    Append a prediction result row to the Prediction_Log sheet.
    Returns True on success.
    """
    if not excel_exists():
        return False
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if "Prediction_Log" not in wb.sheetnames:
            wb.close()
            return False

        ws = wb["Prediction_Log"]

        # Find next empty row (after header rows)
        next_row = ws.max_row + 1
        if next_row < 3:
            next_row = 3

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        def s(style="thin", color="1E3050"):
            sd = Side(style=style, color=color)
            return Border(left=sd, right=sd, top=sd, bottom=sd)

        bg = "0D1A2E" if (next_row % 2 == 0) else "091422"
        risk_color_map = {
            "Extreme": "FF4757", "High": "FF9A3C",
            "Moderate": "FFD166", "Low": "00D4AA"
        }
        fg = risk_color_map.get(result_dict.get("risk_level", "Low"), "C8D8F0")

        row_data = [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            round(result_dict.get("cloudburst_probability", 0), 1),
            result_dict.get("risk_level", "—"),
            round(result_dict.get("rainfall_intensity", 0), 1),
            round(result_dict.get("flash_flood_probability", 0), 1),
            round(result_dict.get("landslide_probability", 0), 1),
            result_dict.get("early_warning_hours", 3),
            result_dict.get("alert_level", "NORMAL"),
            "; ".join(result_dict.get("recommendations", [])[:2]),
        ]

        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=next_row, column=col)
            c.value = val
            c.font  = Font(name="Arial", size=9, color=fg)
            c.fill  = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = s()

        ws.row_dimensions[next_row].height = 18
        wb.save(EXCEL_FILE)
        wb.close()
        return True
    except Exception:
        return False
